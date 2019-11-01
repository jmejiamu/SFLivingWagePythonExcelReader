# import csv
import pymysql.cursors
import time
import openpyxl

#Connecting to the database
try:
    connection = pymysql.connect(
                                    host='xxxxx',    #any IP or localhost or hostname
                                    user= 'xxxx',       #username
                                    password = 'xxxx',      #password
                                    db='xxxxxxx', #DB name
                                    charset='utf8',
                                    cursorclass= pymysql.cursors.DictCursor
                                )
    print("Connected....")
except pymysql.err.OperationalError as error:
    print("User name or password are incorrect!")
    print("Try again!")
    exit()
  
# Give the location of the file 
path = "C:\\Users\\joch_\\Desktop\\Pyhton\\events.xlsx"

# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
sheet_obj = wb_obj.active 
max_col = sheet_obj.max_column 
data = []
# Loop will print all columns name 
for i in range(1, max_col + 1): 
    cell_obj = sheet_obj.cell(row = 2, column = i) 
    data.append(cell_obj.value)
 

#Inserting data to the database
try:
    cursor = connection.cursor()
    sql = ("INSERT INTO calendar VALUES('{}' ,'{}', '{}', '{}', '{}','{}','{}')"
                                                .format(data[0], data[1], data[2], data[3], data[4], data[5], data[6]))
    cursor.execute(sql)
    connection.commit()
    connection.close() #close connection
except pymysql.err.IntegrityError:
    print("Change the id number in the events.xlsx")